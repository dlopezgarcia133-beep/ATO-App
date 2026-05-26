from collections import defaultdict
from io import BytesIO
from typing import List, Optional
import re
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app import models, schemas
from app.config import get_current_user
from app.database import get_db
from app.services import obtener_desglose_comisiones_por_empleado

router = APIRouter()

FORMAS_PAGO_VALIDAS = {"BBVA", "Banco Azteca", "Kids"}


class NominaGrupoUpdate(BaseModel):
    sueldo_base: float
    forma_pago: Optional[str] = None
    cuenta_clabe: Optional[str] = None
    cuenta_interbancaria: Optional[str] = None


class EnglobadoUpdate(BaseModel):
    nombre_englobado: Optional[str] = None


class JornadaFijaUpdate(BaseModel):
    jornada: float


class DiaTrabajo(BaseModel):
    dia: str
    entrada: Optional[str] = None
    salida: Optional[str] = None
    descanso: bool = False


class HorarioUpdate(BaseModel):
    horario_semanal: List[DiaTrabajo]
    dia_descanso: Optional[str] = None


def _calcular_horas(entrada: str, salida: str) -> float:
    """Misma fórmula que frontend: diff en minutos / 60."""
    eh, em = map(int, entrada.split(":"))
    sh, sm = map(int, salida.split(":"))
    diff_min = (sh * 60 + sm) - (eh * 60 + em)
    return 0.0 if diff_min <= 0 else diff_min / 60


def _solo_admin(user: models.Usuario) -> None:
    if user.rol != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")


@router.get("/nomina")
def get_nomina_consolidada(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    usuarios = db.query(models.Usuario).filter(models.Usuario.activo == True).all()

    groups: dict = defaultdict(list)
    for u in usuarios:
        key = u.nombre_englobado or u.username
        groups[key].append(u)

    result = []
    for group_name, perfiles in sorted(groups.items()):
        principal = next((p for p in perfiles if (p.sueldo_base or 0) > 0), perfiles[0])
        sueldo_total = sum(p.sueldo_base or 0 for p in perfiles)

        result.append({
            "nombre_englobado": group_name,
            "sueldo_base_total": sueldo_total,
            "forma_pago": principal.forma_pago,
            "cuenta_clabe": principal.cuenta_clabe,
            "cuenta_interbancaria": principal.cuenta_interbancaria,
            "perfiles_incluidos": [
                {
                    "id": p.id,
                    "codigo": p.username,
                    "modulo": p.modulo.nombre if p.modulo else None,
                    "sueldo": p.sueldo_base or 0,
                }
                for p in sorted(perfiles, key=lambda x: x.id)
            ],
        })

    return result


@router.put("/nomina/{nombre_englobado}")
def update_nomina_grupo(
    nombre_englobado: str,
    data: NominaGrupoUpdate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    if data.forma_pago and data.forma_pago not in FORMAS_PAGO_VALIDAS:
        raise HTTPException(400, f"forma_pago debe ser uno de: {', '.join(FORMAS_PAGO_VALIDAS)}")
    if data.sueldo_base < 0:
        raise HTTPException(400, "sueldo_base debe ser >= 0")
    if data.cuenta_clabe and not (18 <= len(data.cuenta_clabe) <= 20):
        raise HTTPException(400, "cuenta_clabe debe tener entre 18 y 20 caracteres")

    # Buscar por nombre_englobado
    perfiles = (
        db.query(models.Usuario)
        .filter(
            models.Usuario.nombre_englobado == nombre_englobado,
            models.Usuario.activo == True,
        )
        .all()
    )

    # Si no hay, asumir grupo de uno (clave == username, sin nombre_englobado)
    if not perfiles:
        perfiles = (
            db.query(models.Usuario)
            .filter(
                models.Usuario.username == nombre_englobado,
                models.Usuario.activo == True,
                models.Usuario.nombre_englobado.is_(None),
            )
            .all()
        )

    if not perfiles:
        raise HTTPException(404, "Grupo no encontrado")

    principal = next((p for p in perfiles if (p.sueldo_base or 0) > 0), perfiles[0])

    for p in perfiles:
        p.forma_pago = data.forma_pago or None
        p.cuenta_clabe = data.cuenta_clabe or None
        p.cuenta_interbancaria = data.cuenta_interbancaria or None
        p.sueldo_base = data.sueldo_base if p.id == principal.id else 0

    db.commit()
    return {"ok": True}


@router.put("/usuarios/{usuario_id}/englobado")
def assign_nombre_englobado(
    usuario_id: int,
    data: EnglobadoUpdate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    usuario = db.query(models.Usuario).filter_by(id=usuario_id).first()
    if not usuario:
        raise HTTPException(404, "Usuario no encontrado")

    usuario.nombre_englobado = data.nombre_englobado or None
    db.commit()
    return {"ok": True}


@router.put("/usuarios/{usuario_id}/jornada-fija")
def update_jornada_fija(
    usuario_id: int,
    data: JornadaFijaUpdate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    usuario = db.query(models.Usuario).filter_by(id=usuario_id).first()
    if not usuario:
        raise HTTPException(404, "Usuario no encontrado")

    usuario.jornada_fija = data.jornada
    db.commit()
    return {"ok": True}


@router.put("/usuarios/{usuario_id}/horario")
def update_horario(
    usuario_id: int,
    data: HorarioUpdate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    usuario = db.query(models.Usuario).filter_by(id=usuario_id).first()
    if not usuario:
        raise HTTPException(404, "Usuario no encontrado")

    jornada_total = sum(
        _calcular_horas(d.entrada, d.salida)
        for d in data.horario_semanal
        if not d.descanso and d.entrada and d.salida
    )
    jornada_total = round(jornada_total, 2)

    usuario.horario_semanal = [d.model_dump() for d in data.horario_semanal]
    usuario.dia_descanso = data.dia_descanso or None
    usuario.jornada_fija = jornada_total

    db.commit()
    return {"ok": True, "jornada_fija": jornada_total}


@router.post("/ciclos-guardados", response_model=schemas.CicloGuardadoResponse)
def crear_ciclo_guardado(
    data: schemas.CicloGuardadoCreate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    if not data.etiqueta.strip():
        raise HTTPException(400, "La etiqueta no puede estar vacía")

    registro = models.CicloGuardado(
        concepto=data.concepto,
        etiqueta=data.etiqueta.strip(),
        fecha_inicio=data.fecha_inicio,
        fecha_fin=data.fecha_fin,
        datos=data.datos,
        creado_por=current_user.username,
    )
    db.add(registro)
    db.commit()
    db.refresh(registro)
    return registro


@router.get("/ciclos-guardados", response_model=list[schemas.CicloGuardadoResponse])
def listar_ciclos_guardados(
    concepto: str = Query(...),
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    return (
        db.query(models.CicloGuardado)
        .filter(models.CicloGuardado.concepto == concepto)
        .order_by(models.CicloGuardado.creado_en.desc())
        .all()
    )


# ── Comisiones por grupo ──────────────────────────────────────────────────────

_GRUPOS_CONFIG = {
    "asesores":   {"rol": models.RolEnum.asesor,    "pattern": "^A[0-9]", "seccion": "comisiones_asesores"},
    "encargados": {"rol": models.RolEnum.encargado,  "pattern": "^A[0-9]", "seccion": "comisiones_encargados"},
    "cadenas":    {"rol": models.RolEnum.asesor,     "pattern": "^C[0-9]", "seccion": "comisiones_cadenas"},
}


@router.get("/comisiones-por-grupo")
def get_comisiones_por_grupo(
    grupo: str = Query(...),
    fecha_inicio: date = Query(...),
    fecha_fin: date = Query(...),
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    if grupo not in _GRUPOS_CONFIG:
        raise HTTPException(400, f"grupo debe ser uno de: {', '.join(_GRUPOS_CONFIG)}")

    cfg = _GRUPOS_CONFIG[grupo]

    usuarios = (
        db.query(models.Usuario)
        .filter(
            models.Usuario.activo == True,
            models.Usuario.rol == cfg["rol"],
            models.Usuario.username.op("~*")(cfg["pattern"]),
        )
        .all()
    )

    desglose_map = obtener_desglose_comisiones_por_empleado(db, fecha_inicio, fecha_fin)

    groups: dict = defaultdict(list)
    for u in usuarios:
        key = u.nombre_englobado or u.username
        groups[key].append(u)

    result = []
    for group_name, perfiles in sorted(groups.items()):
        acc   = sum(desglose_map.get(p.id, {}).get("accesorios", 0) for p in perfiles)
        tel   = sum(desglose_map.get(p.id, {}).get("telefonos",  0) for p in perfiles)
        chips = sum(desglose_map.get(p.id, {}).get("chips",      0) for p in perfiles)
        total = acc + tel + chips

        nombre = next(
            (p.nombre_completo for p in perfiles if p.nombre_completo),
            group_name,
        )

        result.append({
            "seccion": cfg["seccion"],
            "empleado": group_name,
            "nombre_completo": nombre,
            "comisiones_accesorios": round(acc, 2),
            "comisiones_telefonos": round(tel, 2),
            "comisiones_chips": round(chips, 2),
            "comisiones_total": round(total, 2),
            "pago_total": round(total, 2),
            "tiene_comisiones": total > 0,
        })

    return result


# ── Nóminas ───────────────────────────────────────────────────────────────────

def _safe_filename(text: str) -> str:
    return re.sub(r"[^\w\-]", "_", text)[:60]


@router.post("/nominas", response_model=schemas.NominaResponse)
def crear_nomina(
    data: schemas.NominaCreate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    if not data.etiqueta.strip():
        raise HTTPException(400, "La etiqueta no puede estar vacía")

    total = sum(float(item.get("pago_total", 0)) for item in data.datos)

    nomina = models.Nomina(
        etiqueta=data.etiqueta.strip(),
        ciclo_horas_extras_id=data.ciclo_horas_extras_id,
        fecha_inicio_asesores=data.fecha_inicio_asesores,
        fecha_fin_asesores=data.fecha_fin_asesores,
        fecha_inicio_encargados=data.fecha_inicio_encargados,
        fecha_fin_encargados=data.fecha_fin_encargados,
        fecha_inicio_cadenas=data.fecha_inicio_cadenas,
        fecha_fin_cadenas=data.fecha_fin_cadenas,
        total_pago=round(total, 2),
        datos=data.datos,
        creado_por=current_user.username,
    )
    db.add(nomina)
    db.commit()
    db.refresh(nomina)
    return nomina


@router.get("/nominas", response_model=list[schemas.NominaListItem])
def listar_nominas(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    return (
        db.query(models.Nomina)
        .order_by(models.Nomina.creado_en.desc())
        .all()
    )


@router.get("/nominas/{nomina_id}", response_model=schemas.NominaResponse)
def detalle_nomina(
    nomina_id: int,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    nomina = db.query(models.Nomina).filter_by(id=nomina_id).first()
    if not nomina:
        raise HTTPException(404, "Nómina no encontrada")
    return nomina


def _excel_sheet_horas_extras(wb, datos, header_font, header_fill, total_font):
    from openpyxl.styles import Alignment
    ws = wb.active
    ws.title = "Horas Extras"
    headers = ["Empleado", "Nombre completo", "H. Extra", "Pago H. Extras", "Total Pago"]
    col_widths = [20, 30, 14, 18, 18]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w
    rows = [d for d in datos if d.get("seccion") == "horas_extras"]
    for row_idx, item in enumerate(rows, start=2):
        he = item.get("horas_extra_redondeo")
        he_str = "—" if he is None else (f"+{he}h" if float(he) > 0 else f"{he}h")
        ws.cell(row=row_idx, column=1, value=item.get("empleado", ""))
        ws.cell(row=row_idx, column=2, value=item.get("nombre_completo", ""))
        ws.cell(row=row_idx, column=3, value=he_str)
        ws.cell(row=row_idx, column=4, value=round(float(item.get("pago_horas_extras", 0)), 2))
        ws.cell(row=row_idx, column=5, value=round(float(item.get("pago_total", 0)), 2))
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=5, value=round(sum(float(d.get("pago_total", 0)) for d in rows), 2)).font = total_font


def _excel_sheet_comisiones(wb, titulo, seccion_key, datos, header_font, header_fill, total_font):
    from openpyxl.styles import Alignment
    ws = wb.create_sheet(title=titulo)
    headers = ["Empleado", "Nombre completo", "Accesorios", "Teléfonos", "Chips", "Total"]
    col_widths = [20, 30, 14, 14, 14, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w
    rows = [d for d in datos if d.get("seccion") == seccion_key]
    for row_idx, item in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=item.get("empleado", ""))
        ws.cell(row=row_idx, column=2, value=item.get("nombre_completo", ""))
        ws.cell(row=row_idx, column=3, value=round(float(item.get("comisiones_accesorios", 0)), 2))
        ws.cell(row=row_idx, column=4, value=round(float(item.get("comisiones_telefonos", 0)), 2))
        ws.cell(row=row_idx, column=5, value=round(float(item.get("comisiones_chips", 0)), 2))
        ws.cell(row=row_idx, column=6, value=round(float(item.get("pago_total", 0)), 2))
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=6, value=round(sum(float(d.get("pago_total", 0)) for d in rows), 2)).font = total_font


@router.get("/nominas/{nomina_id}/excel")
def descargar_nomina_excel(
    nomina_id: int,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    nomina = db.query(models.Nomina).filter_by(id=nomina_id).first()
    if not nomina:
        raise HTTPException(404, "Nómina no encontrada")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="FF6600")
    total_font = Font(bold=True)

    wb = Workbook()

    secciones_presentes = {d.get("seccion") for d in nomina.datos}

    _excel_sheet_horas_extras(wb, nomina.datos, header_font, header_fill, total_font)

    if "comisiones_asesores" in secciones_presentes:
        _excel_sheet_comisiones(wb, "Comisiones Asesores", "comisiones_asesores", nomina.datos, header_font, header_fill, total_font)
    if "comisiones_encargados" in secciones_presentes:
        _excel_sheet_comisiones(wb, "Comisiones Encargados", "comisiones_encargados", nomina.datos, header_font, header_fill, total_font)
    if "comisiones_cadenas" in secciones_presentes:
        _excel_sheet_comisiones(wb, "Comisiones Cadenas", "comisiones_cadenas", nomina.datos, header_font, header_fill, total_font)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"nomina_{_safe_filename(nomina.etiqueta)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _pdf_table_horas_extras(datos, colors, ato_orange):
    from reportlab.platypus import Table, TableStyle
    rows = [d for d in datos if d.get("seccion") == "horas_extras"]
    table_data = [["Empleado", "Nombre completo", "H. Extra", "Pago H. Extras", "Total Pago"]]
    for item in rows:
        he = item.get("horas_extra_redondeo")
        he_str = "—" if he is None else (f"+{he}h" if float(he) > 0 else f"{he}h")
        table_data.append([
            item.get("empleado", ""),
            item.get("nombre_completo", ""),
            he_str,
            f"${float(item.get('pago_horas_extras', 0)):.2f}",
            f"${float(item.get('pago_total', 0)):.2f}",
        ])
    subtotal = sum(float(d.get("pago_total", 0)) for d in rows)
    table_data.append(["TOTAL", "", "", "", f"${subtotal:.2f}"])
    t = Table(table_data, colWidths=[90, 155, 65, 90, 90])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ato_orange),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return t


def _pdf_table_comisiones(seccion_key, datos, colors, ato_orange):
    from reportlab.platypus import Table, TableStyle
    rows = [d for d in datos if d.get("seccion") == seccion_key]
    table_data = [["Empleado", "Nombre completo", "Accesorios", "Teléfonos", "Chips", "Total"]]
    for item in rows:
        table_data.append([
            item.get("empleado", ""),
            item.get("nombre_completo", ""),
            f"${float(item.get('comisiones_accesorios', 0)):.2f}",
            f"${float(item.get('comisiones_telefonos', 0)):.2f}",
            f"${float(item.get('comisiones_chips', 0)):.2f}",
            f"${float(item.get('pago_total', 0)):.2f}",
        ])
    subtotal = sum(float(d.get("pago_total", 0)) for d in rows)
    table_data.append(["TOTAL", "", "", "", "", f"${subtotal:.2f}"])
    t = Table(table_data, colWidths=[80, 130, 65, 65, 60, 65])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ato_orange),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return t


@router.get("/nominas/{nomina_id}/pdf")
def descargar_nomina_pdf(
    nomina_id: int,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    nomina = db.query(models.Nomina).filter_by(id=nomina_id).first()
    if not nomina:
        raise HTTPException(404, "Nómina no encontrada")

    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    ato_orange = colors.HexColor("#FF6600")

    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=ato_orange, spaceAfter=2)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=ato_orange, fontSize=11, spaceAfter=4)

    elements = []
    elements.append(Paragraph(f"NÓMINA — {nomina.etiqueta}", h1))
    creado_str = nomina.creado_en.strftime("%d/%m/%Y %H:%M") if nomina.creado_en else ""
    elements.append(Paragraph(f"Creado por: {nomina.creado_por}  |  {creado_str}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    secciones_presentes = {d.get("seccion") for d in nomina.datos}

    elements.append(Paragraph("Horas Extras", h2))
    elements.append(_pdf_table_horas_extras(nomina.datos, colors, ato_orange))

    secciones_comisiones = [
        ("comisiones_asesores",   "Comisiones Asesores"),
        ("comisiones_encargados", "Comisiones Encargados"),
        ("comisiones_cadenas",    "Comisiones Cadenas"),
    ]
    for key, titulo in secciones_comisiones:
        if key in secciones_presentes:
            elements.append(Spacer(1, 10))
            elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
            elements.append(Spacer(1, 6))
            elements.append(Paragraph(titulo, h2))
            elements.append(_pdf_table_comisiones(key, nomina.datos, colors, ato_orange))

    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<b>TOTAL NÓMINA: ${float(nomina.total_pago):.2f}</b>", styles["Normal"]))

    doc.build(elements)
    buf.seek(0)

    fname = f"nomina_{_safe_filename(nomina.etiqueta)}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Nóminas de Incubadora ─────────────────────────────────────────────────────

@router.get("/chips-incubadora-pendientes")
def get_chips_incubadora_pendientes(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    chips = (
        db.query(models.VentaChip)
        .filter(
            models.VentaChip.es_incubadora == True,
            models.VentaChip.validado == True,
            models.VentaChip.comision_pagada == False,
            models.VentaChip.cancelada == False,
            models.VentaChip.numero_telefono.isnot(None),
        )
        .all()
    )

    # Agrupar por empleado_id para luego colapsar por nombre_englobado
    by_user: dict = defaultdict(list)
    for c in chips:
        by_user[c.empleado_id].append(c)

    # Cargar usuarios relevantes en un solo query
    user_ids = list(by_user.keys())
    usuarios = {
        u.id: u
        for u in db.query(models.Usuario).filter(models.Usuario.id.in_(user_ids)).all()
    }

    # Agrupar por nombre_englobado
    groups: dict = defaultdict(lambda: {"perfiles": [], "chips": []})
    for uid, chip_list in by_user.items():
        u = usuarios.get(uid)
        if not u:
            continue
        key = u.nombre_englobado or u.username
        groups[key]["perfiles"].append(u)
        groups[key]["chips"].extend(chip_list)

    result = []
    for group_name, gdata in sorted(groups.items()):
        perfiles = gdata["perfiles"]
        group_chips = gdata["chips"]
        total = sum(float(c.comision or 0) for c in group_chips)
        nombre = next((p.nombre_completo for p in perfiles if p.nombre_completo), group_name)
        usuario_ids = sorted({p.id for p in perfiles})

        result.append({
            "empleado": group_name,
            "nombre_completo": nombre,
            "usuario_ids": usuario_ids,
            "chips_count": len(group_chips),
            "total_chips_incubadora": round(total, 2),
            "pago_total": round(total, 2),
            "detalle": [
                {
                    "chip_id": c.id,
                    "tipo_chip": c.tipo_chip,
                    "numero_telefono": c.numero_telefono,
                    "comision": round(float(c.comision or 0), 2),
                    "fecha_venta": str(c.fecha),
                }
                for c in sorted(group_chips, key=lambda x: x.fecha)
            ],
        })

    return result


@router.post("/nominas-incubadora", response_model=schemas.NominaIncubadoraResponse)
def crear_nomina_incubadora(
    data: schemas.NominaIncubadoraCreate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)

    if not data.etiqueta.strip():
        raise HTTPException(400, "La etiqueta no puede estar vacía")
    if not data.datos:
        raise HTTPException(400, "La nómina no tiene empleados")

    # Extraer todos los chip_ids del snapshot
    chip_ids = [d["chip_id"] for item in data.datos for d in item.get("detalle", [])]
    if not chip_ids:
        raise HTTPException(400, "No hay chips en el detalle de la nómina")

    # Guard anti-duplicado: verificar que ningún chip ya esté pagado
    ya_pagados = (
        db.query(models.VentaChip.id)
        .filter(
            models.VentaChip.id.in_(chip_ids),
            models.VentaChip.comision_pagada == True,
        )
        .all()
    )
    if ya_pagados:
        ids_str = ", ".join(str(r.id) for r in ya_pagados)
        raise HTTPException(409, f"Los chips {ids_str} ya fueron pagados en una nómina anterior")

    total = sum(float(item.get("pago_total", 0)) for item in data.datos)

    try:
        nomina = models.NominaIncubadora(
            etiqueta=data.etiqueta.strip(),
            total_pago=round(total, 2),
            datos=data.datos,
            creado_por=current_user.username,
        )
        db.add(nomina)
        db.flush()  # obtener el id antes del commit

        db.query(models.VentaChip).filter(
            models.VentaChip.id.in_(chip_ids),
            models.VentaChip.comision_pagada == False,
        ).update({"comision_pagada": True}, synchronize_session=False)

        db.commit()
        db.refresh(nomina)
        return nomina
    except Exception:
        db.rollback()
        raise HTTPException(500, "Error al guardar la nómina; se revirtió la transacción")


@router.get("/nominas-incubadora", response_model=list[schemas.NominaIncubadoraListItem])
def listar_nominas_incubadora(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)
    return (
        db.query(models.NominaIncubadora)
        .order_by(models.NominaIncubadora.creado_en.desc())
        .all()
    )


@router.get("/nominas-incubadora/{nomina_id}", response_model=schemas.NominaIncubadoraResponse)
def detalle_nomina_incubadora(
    nomina_id: int,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)
    nomina = db.query(models.NominaIncubadora).filter_by(id=nomina_id).first()
    if not nomina:
        raise HTTPException(404, "Nómina de incubadora no encontrada")
    return nomina


@router.get("/nominas-incubadora/{nomina_id}/excel")
def descargar_nomina_incubadora_excel(
    nomina_id: int,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)
    nomina = db.query(models.NominaIncubadora).filter_by(id=nomina_id).first()
    if not nomina:
        raise HTTPException(404, "Nómina de incubadora no encontrada")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="7C3AED")
    tf = Font(bold=True)
    center = Alignment(horizontal="center")

    # ── Hoja 1: Resumen ──
    ws1 = wb.active
    ws1.title = "Resumen"
    for col, (h, w) in enumerate(
        zip(["Empleado", "Nombre completo", "N° Chips", "Total"], [22, 32, 12, 16]),
        start=1,
    ):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = center
        ws1.column_dimensions[cell.column_letter].width = w

    for i, item in enumerate(nomina.datos, start=2):
        ws1.cell(row=i, column=1, value=item.get("empleado", ""))
        ws1.cell(row=i, column=2, value=item.get("nombre_completo", ""))
        ws1.cell(row=i, column=3, value=item.get("chips_count", 0))
        ws1.cell(row=i, column=4, value=round(float(item.get("pago_total", 0)), 2))

    total_row = len(nomina.datos) + 2
    ws1.cell(row=total_row, column=1, value="TOTAL").font = tf
    ws1.cell(row=total_row, column=4, value=float(nomina.total_pago)).font = tf

    # ── Hoja 2: Detalle ──
    ws2 = wb.create_sheet(title="Detalle")
    det_headers = ["Empleado", "Nombre completo", "Tipo Chip", "Número", "Comisión", "Fecha Venta"]
    det_widths  = [22, 32, 18, 16, 14, 14]
    for col, (h, w) in enumerate(zip(det_headers, det_widths), start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = center
        ws2.column_dimensions[cell.column_letter].width = w

    row_idx = 2
    for item in nomina.datos:
        emp = item.get("empleado", "")
        nombre = item.get("nombre_completo", "")
        for chip in item.get("detalle", []):
            ws2.cell(row=row_idx, column=1, value=emp)
            ws2.cell(row=row_idx, column=2, value=nombre)
            ws2.cell(row=row_idx, column=3, value=chip.get("tipo_chip", ""))
            ws2.cell(row=row_idx, column=4, value=chip.get("numero_telefono", ""))
            ws2.cell(row=row_idx, column=5, value=round(float(chip.get("comision", 0)), 2))
            ws2.cell(row=row_idx, column=6, value=chip.get("fecha_venta", ""))
            row_idx += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"incubadora_{_safe_filename(nomina.etiqueta)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/nominas-incubadora/{nomina_id}/pdf")
def descargar_nomina_incubadora_pdf(
    nomina_id: int,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _solo_admin(current_user)
    nomina = db.query(models.NominaIncubadora).filter_by(id=nomina_id).first()
    if not nomina:
        raise HTTPException(404, "Nómina de incubadora no encontrada")

    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    purple = colors.HexColor("#7C3AED")

    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=purple, spaceAfter=2)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=purple, fontSize=11, spaceAfter=4)

    elements = []
    elements.append(Paragraph(f"NÓMINA DE INCUBADORA — {nomina.etiqueta}", h1))
    creado_str = nomina.creado_en.strftime("%d/%m/%Y %H:%M") if nomina.creado_en else ""
    elements.append(Paragraph(f"Creado por: {nomina.creado_por}  |  {creado_str}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    # ── Resumen ──
    elements.append(Paragraph("Resumen", h2))
    res_data = [["Empleado", "Nombre completo", "N° Chips", "Total"]]
    for item in nomina.datos:
        res_data.append([
            item.get("empleado", ""),
            item.get("nombre_completo", ""),
            str(item.get("chips_count", 0)),
            f"${float(item.get('pago_total', 0)):.2f}",
        ])
    res_data.append(["TOTAL", "", "", f"${float(nomina.total_pago):.2f}"])

    t_res = Table(res_data, colWidths=[90, 160, 70, 80])
    t_res.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), purple),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f5f3ff")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(t_res)

    # ── Detalle por empleado ──
    elements.append(Spacer(1, 12))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph("Detalle por Empleado", h2))

    for item in nomina.datos:
        if not item.get("detalle"):
            continue
        elements.append(Paragraph(
            f"<b>{item.get('empleado','')} — {item.get('nombre_completo','')}</b>",
            styles["Normal"],
        ))
        elements.append(Spacer(1, 3))

        det_data = [["Tipo Chip", "Número", "Comisión", "Fecha Venta"]]
        for chip in item.get("detalle", []):
            det_data.append([
                chip.get("tipo_chip", ""),
                chip.get("numero_telefono", ""),
                f"${float(chip.get('comision', 0)):.2f}",
                chip.get("fecha_venta", ""),
            ])

        t_det = Table(det_data, colWidths=[110, 120, 80, 90])
        t_det.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ede9fe")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(t_det)
        elements.append(Spacer(1, 6))

    doc.build(elements)
    buf.seek(0)

    fname = f"incubadora_{_safe_filename(nomina.etiqueta)}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
